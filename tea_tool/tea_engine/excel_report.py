"""Comprehensive multi-sheet Excel report generator.

Produces ONE workbook with everything a reviewer needs to spot-check the TEA:

  - Read_me            sheet index + what to look at + how to verify
  - PFD                embedded process flow diagram (PNG)
  - Process_structure  source-of-truth tables (sections, edges, streams,
                       equipment) — verify these match your understanding
  - Summary            top-line economics per scale
  - CAPEX_breakdown    equipment cost per section × scale
  - OPEX_breakdown     OPEX line items × scale
  - Revenue_breakdown  output revenue + market prices
  - Flows_annual       annual mass flows per component
  - Levers             every build() lever + lab/commercial reference
  - Gap_report         lever ranking by economic impact (breakeven solver)
  - Sweep_<lever>      one sheet per lever — see the sensitivity curve
  - Price_<lev_x>_<lev_y>  2-D price grid for selected price levers
  - Components         ComponentDB dump for full transparency

Every data sheet starts with a 3-line yellow header that says:
  (1) what this sheet shows
  (2) units / how to read it
  (3) what to spot-check for errors

Usage:
    from processes.spent_lfp_ballmill_li import build
    from tea_engine.excel_report import full_report
    full_report(build, output_path="out/lfp_full_report.xlsx")
"""
from __future__ import annotations
import importlib
import os
import tempfile
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from .tea import run_tea
from .solver import breakeven


# ----------------------------------------------------------------------------
# Cell formatting helpers
# ----------------------------------------------------------------------------

def _ensure_dir(path: str) -> None:
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def _safe_sheet_name(name: str) -> str:
    bad = '/\\?*[]:'
    s = name
    for ch in bad:
        s = s.replace(ch, "_")
    return s[:31]


def _bold():
    from openpyxl.styles import Font
    return Font(bold=True)


def _bold_italic():
    from openpyxl.styles import Font
    return Font(bold=True, italic=True)


def _italic():
    from openpyxl.styles import Font
    return Font(italic=True)


def _header_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="D9D9D9")


def _note_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="FFF9C4")  # light yellow


def _money_fmt(cell):
    cell.number_format = '"$"#,##0'


def _money_fmt2(cell):
    cell.number_format = '"$"#,##0.0000'


def _num_fmt(cell):
    cell.number_format = '#,##0.0000'


def _write_note_block(ws, lines: List[str], start_row: int = 1) -> int:
    """Write yellow-shaded info block. Returns next free row."""
    for i, line in enumerate(lines):
        c = ws.cell(start_row + i, 1, line)
        c.font = _italic()
        c.fill = _note_fill()
        # Merge across columns A-H so the note spans wide
        ws.merge_cells(start_row=start_row + i, end_row=start_row + i,
                       start_column=1, end_column=8)
    return start_row + len(lines) + 1   # +1 = blank row separator


# ----------------------------------------------------------------------------
# Read_me sheet
# ----------------------------------------------------------------------------

def _write_readme(wb, build_fn, inp) -> None:
    ws = wb.create_sheet("Read_me", 0)
    p, db, _ = build_fn()
    rows = [
        ("TEA Report — How to use this workbook", "header"),
        (f"Process: {p.name}", "info"),
        (f"Module:  {build_fn.__module__}", "info"),
        ("", ""),
        ("Reading order (top → bottom):", "section"),
        ("  1. Read_me           This page",                  "row"),
        ("  2. Diagrams          BFD + PFD + P&ID stacked in one sheet (visual verification)", "row"),
        ("  3. Process_structure Sections / edges / streams / equipment tables (verify YOUR model is the one being computed)", "row"),
        ("  4. Summary           Headline $/kg-feed + CAPEX/OPEX/Revenue at every scale", "row"),
        ("  5. CAPEX_breakdown   Where the CAPEX goes (section × scale)", "row"),
        ("  6. OPEX_breakdown    Where the OPEX goes (line item × scale)", "row"),
        ("  7. Revenue_breakdown Which products earn how much (× scale)", "row"),
        ("  8. Flows_annual      Mass balance check (kg-component/year × scale)", "row"),
        ("  9. Levers            Every adjustable parameter + its lab/commercial value", "row"),
        (" 10. Gap_report        Lever ranking — which knob moves the economy most", "row"),
        (" 11. Sweep_<lever>     Per-lever sensitivity curve (one sheet each)", "row"),
        (" 12. Price_<...>       2-D grid showing impact of two market-price levers", "row"),
        (" 13. Components        Every component's MW, price, role — for transparency", "row"),
        ("", ""),
        ("How to verify there is NO error (recommended quick checks):", "section"),
        ("  A. On PFD/Process_structure — every input you expect appears as 'IN', every product as 'OUT', recycle loops are dashed red.", "row"),
        ("  B. Flows_annual — feed kg/y × $/kg should match OPEX line 'Feedstock - <name>'.", "row"),
        ("  C. Revenue_breakdown — product kg/y × $/kg should match Summary's Revenue total.", "row"),
        ("  D. CAPEX_breakdown — section totals × CRF (= discount * 1/(1-(1+d)^-life)) ≈ CAPEX_annualized on Summary.", "row"),
        ("  E. Summary — Revenue - (CAPEX_annualized + OPEX) = Net profit (within rounding).", "row"),
        ("  F. Net profit / feed annual kg = $/kg-feedstock.", "row"),
        ("  G. Gap_report 'Breakeven (alone)' = parameter value that makes $/kg-feedstock = 0 with all other levers at lab.", "row"),
        ("", ""),
        ("Key TEA assumptions (from TEAInputs):", "section"),
        (f"  Discount rate:     {inp.discount_rate}", "row"),
        (f"  Plant life:        {inp.lifetime_years} years", "row"),
        (f"  Capacity factor:   {inp.capacity_factor}", "row"),
        (f"  OSBL fraction:     {inp.osbl_fraction}  (OSBL = OSBL_fraction × ISBL)", "row"),
        (f"  Maintenance:       {inp.maintenance_fraction} × CAPEX/y", "row"),
        (f"  Operation:         {inp.operation_fraction} × CAPEX/y", "row"),
        (f"  Batch hours:       {inp.batch_hours} h/batch  →  {round(8760 * inp.capacity_factor / inp.batch_hours, 0):.0f} batches/year", "row"),
        (f"  CEPCI target year: {inp.cepci_target_year}", "row"),
        (f"  MSP product:       {inp.msp_product}", "row"),
        (f"  Feedstock basis:   {inp.feedstock_for_economics}", "row"),
        (f"  Scales evaluated:  {list(inp.scales_ton)} t-feed/batch", "row"),
        ("", ""),
        ("If any number seems wrong — first compare Process_structure + Flows_annual.", "section"),
        ("Then trace from there to OPEX/Revenue/CAPEX. The Read_me checks above are sufficient to spot 95% of bugs.", "row"),
    ]

    r = 1
    for text, kind in rows:
        c = ws.cell(r, 1, text)
        if kind == "header":
            from openpyxl.styles import Font
            c.font = Font(bold=True, size=14)
        elif kind == "section":
            c.font = _bold_italic()
        elif kind == "info":
            c.font = _italic()
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=8)
        r += 1

    ws.column_dimensions["A"].width = 110


# ----------------------------------------------------------------------------
# Diagrams sheet (BFD + PFD + P&ID stacked)
# ----------------------------------------------------------------------------

def _write_diagrams(wb, process, bfd_png: str, pfd_png: str, pid_png: str) -> None:
    from openpyxl.drawing.image import Image
    ws = wb.create_sheet("Diagrams")
    r = _write_note_block(ws, [
        "Three diagram tiers (chemical-engineering convention) stacked top → bottom:",
        "  1. BFD  (Block Flow Diagram)    — collapsed unit ops, main flow only. For executive orientation.",
        "  2. PFD  (Process Flow Diagram)  — every section, all streams, recycle loops, stream labels. Engineering-comm level.",
        "  3. P&ID (Piping & Instrumentation Diagram, schematic) — PFD + pumps + control valves + instrument balloons (TI/PI/FI/LI/etc.). Treat as illustrative — instrumentation added heuristically per unit kind.",
        "VERIFY: each diagram should depict the SAME plant. If a section or stream is missing in one but present in another, the underlying Process is inconsistent — fix sections/edges in build().",
    ])

    def _embed(label: str, png: str, row_anchor: int) -> int:
        cell = ws.cell(row_anchor, 1, label)
        cell.font = _bold()
        from openpyxl.styles import Font
        cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row_anchor, end_row=row_anchor,
                       start_column=1, end_column=8)
        img = Image(png)
        # Approximate: control size so each image takes ~28 rows
        img.width = 1100
        img.height = 380  # ish
        img.anchor = ws.cell(row_anchor + 1, 1).coordinate
        ws.add_image(img)
        return row_anchor + 28

    r = _embed("1. BFD — Block Flow Diagram (simplified main flow)", bfd_png, r)
    r = _embed("2. PFD — Process Flow Diagram (sections, streams, recycle)",
               pfd_png, r)
    r = _embed("3. P&ID — Piping & Instrumentation Diagram (schematic)",
               pid_png, r)

    ws.column_dimensions["A"].width = 80


# ----------------------------------------------------------------------------
# Process_structure sheet
# ----------------------------------------------------------------------------

def _write_process_structure(wb, process, db, inp) -> None:
    ws = wb.create_sheet("Process_structure")
    r = _write_note_block(ws, [
        "Source-of-truth tables for the modelled process. Numbers below are what the engine actually computes against.",
        "Tabs in this sheet: Sections | Edges (PFD topology) | Input streams | Output streams | Equipment list",
        "VERIFY: section count matches your design; every stream's mass_per_batch and recovery are correct; equipment base_cost and scaling_factor reflect your CAPEX assumptions.",
    ])

    # --- Sections ---
    c = ws.cell(r, 1, "SECTIONS"); c.font = _bold()
    r += 1
    headers = ["Key", "Label", "Kind (PFD color/icon)", "Description"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    r += 1
    for sec in process.sections:
        ws.cell(r, 1, sec.key)
        ws.cell(r, 2, sec.label)
        ws.cell(r, 3, sec.kind)
        ws.cell(r, 4, sec.description)
        r += 1
    r += 2

    # --- Edges ---
    c = ws.cell(r, 1, "EDGES (PFD connections)"); c.font = _bold()
    r += 1
    headers = ["From", "To", "Label", "Is recycle?"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    r += 1
    # detect recycle by section declaration order
    sec_order = {s.key: i for i, s in enumerate(process.sections)}
    for src, dst, lbl in process.edges:
        is_recycle = ""
        if src in sec_order and dst in sec_order:
            if sec_order[dst] <= sec_order[src]:
                is_recycle = "RECYCLE"
        ws.cell(r, 1, src)
        ws.cell(r, 2, dst)
        ws.cell(r, 3, lbl)
        c = ws.cell(r, 4, is_recycle)
        if is_recycle:
            c.font = _bold()
        r += 1
    r += 2

    # --- Input streams ---
    c = ws.cell(r, 1, "INPUT STREAMS"); c.font = _bold()
    r += 1
    headers = ["Component", "mass per batch (g)", "Recovery (frac)",
               "Category", "Flow mode", "Initial charge (kg/t-feed)",
               "Replacement (months)", "Note"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    r += 1
    for s in process.streams.inputs:
        ws.cell(r, 1, s.component)
        ws.cell(r, 2, s.mass_per_batch_g)
        ws.cell(r, 3, getattr(s, "recovery", 0.0))
        ws.cell(r, 4, getattr(s, "category", ""))
        ws.cell(r, 5, getattr(s, "flow_mode", "continuous"))
        ws.cell(r, 6, getattr(s, "initial_charge_kg_per_ton", "") or "")
        ws.cell(r, 7, getattr(s, "replacement_interval_months", "") or "")
        ws.cell(r, 8, getattr(s, "note", "") or "")
        r += 1
    r += 2

    # --- Output streams ---
    c = ws.cell(r, 1, "OUTPUT STREAMS"); c.font = _bold()
    r += 1
    headers = ["Component", "mass per batch (g)", "Note"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    r += 1
    for s in process.streams.outputs:
        ws.cell(r, 1, s.component)
        ws.cell(r, 2, s.mass_per_batch_g)
        ws.cell(r, 3, getattr(s, "note", "") or "")
        r += 1
    r += 2

    # --- Equipment list ---
    c = ws.cell(r, 1, "EQUIPMENT LIST"); c.font = _bold()
    r += 1
    headers = ["Name", "Section", "Base cost ($)", "Scaling factor",
               "Cap_ref (t/batch)", "CEPCI ref year", "Lifetime (y)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    r += 1
    for e in process.equipment.items:
        ws.cell(r, 1, e.name)
        ws.cell(r, 2, e.section)
        ws.cell(r, 3, e.base_cost); _money_fmt(ws.cell(r, 3))
        ws.cell(r, 4, e.scaling_factor)
        ws.cell(r, 5, e.cap_ref)
        ws.cell(r, 6, e.cepci_ref)
        ws.cell(r, 7, e.lifetime_years)
        r += 1

    for col, w in [("A", 30), ("B", 30), ("C", 22), ("D", 16),
                    ("E", 18), ("F", 18), ("G", 18), ("H", 50)]:
        ws.column_dimensions[col].width = w


# ----------------------------------------------------------------------------
# Existing sheet writers — each now starts with a note block
# ----------------------------------------------------------------------------

def _write_summary(wb, build_fn, result, inp, db) -> None:
    ws = wb.create_sheet("Summary")
    r = _write_note_block(ws, [
        "Top-line economics for every scale in scales_ton. The headline metric is $/kg-feedstock (forward-design).",
        "Read across each scale row: CAPEX_total → CAPEX_annualized → OPEX → Revenue → Net profit → $/kg-feed → MSP cross-check.",
        "VERIFY: Revenue - (CAPEX_annualized + OPEX) should equal Net profit. Net profit / annual-feedstock-kg = $/kg-feed.",
    ])

    headers = ["Scale (t-feed/batch)", "CAPEX total ($)",
               "CAPEX annualized ($/y)", "OPEX total ($/y)",
               "Revenue total ($/y)", "Net profit ($/y)",
               "$/kg-feedstock", f"MSP {inp.msp_product} ($/kg)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    for i, ton in enumerate(inp.scales_ton, start=r + 1):
        ws.cell(i, 1, ton)
        ws.cell(i, 2, result.capex_total[ton]); _money_fmt(ws.cell(i, 2))
        ws.cell(i, 3, result.capex_annualized[ton]); _money_fmt(ws.cell(i, 3))
        ws.cell(i, 4, result.opex_total[ton]); _money_fmt(ws.cell(i, 4))
        ws.cell(i, 5, result.revenue_total[ton]); _money_fmt(ws.cell(i, 5))
        ws.cell(i, 6, result.net_profit[ton]); _money_fmt(ws.cell(i, 6))
        ws.cell(i, 7, result.net_per_kg_feedstock.get(ton, 0.0))
        _money_fmt2(ws.cell(i, 7))
        ws.cell(i, 8, result.msp.get(ton, 0.0)); _money_fmt2(ws.cell(i, 8))

    for col, w in zip("ABCDEFGH", [22, 16, 22, 16, 18, 18, 18, 22]):
        ws.column_dimensions[col].width = w


def _write_capex(wb, result, inp) -> None:
    ws = wb.create_sheet("CAPEX_breakdown")
    r = _write_note_block(ws, [
        "Installed equipment cost per process section, at each scale (one column per scale).",
        "OSBL = OSBL_fraction × (sum of all section ISBL). TOTAL = ISBL + OSBL.",
        "VERIFY: section totals match your equipment list (see Process_structure → EQUIPMENT LIST). Scale ratios should follow 0.6 power-law unless equipment is set to scaling_factor=1.0.",
    ])
    scales = list(inp.scales_ton)
    sections = sorted({sec for ton in scales
                       for sec in result.capex_section.get(ton, {}).keys()})
    ws.cell(r, 1, "Section / OSBL"); ws.cell(r, 1).font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(r, j, f"{ton} t-feed/batch")
        c.font = _bold(); c.fill = _header_fill()
    for i, sec in enumerate(sections, start=r + 1):
        ws.cell(i, 1, sec)
        for j, ton in enumerate(scales, start=2):
            v = result.capex_section.get(ton, {}).get(sec, 0.0)
            ws.cell(i, j, v); _money_fmt(ws.cell(i, j))
    total_row = r + 1 + len(sections)
    c = ws.cell(total_row, 1, "TOTAL (ISBL + OSBL)"); c.font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(total_row, j, result.capex_total[ton])
        c.font = _bold(); _money_fmt(c)
    ws.column_dimensions["A"].width = 42
    for j in range(2, len(scales) + 2):
        ws.column_dimensions[chr(64 + j)].width = 22


def _write_opex(wb, result, inp) -> None:
    ws = wb.create_sheet("OPEX_breakdown")
    r = _write_note_block(ws, [
        "Annual OPEX line items at each scale.",
        "Feedstock costs = annual mass × $/kg (from Components sheet). Maintenance/Operation = fraction × CAPEX_total. Periodic costs amortized to /y.",
        "VERIFY: 'Feedstock - X' line = (kg-X/y on Flows_annual) × (price on Components). Maintenance ≈ inp.maintenance_fraction × CAPEX. TOTAL OPEX should match Summary's OPEX_total.",
    ])
    scales = list(inp.scales_ton)
    line_items = sorted({k for ton in scales
                         for k in result.opex.get(ton, {}).keys()
                         if not k.startswith("__")})
    ws.cell(r, 1, "OPEX line item"); ws.cell(r, 1).font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(r, j, f"{ton} t-feed/batch ($/y)")
        c.font = _bold(); c.fill = _header_fill()
    for i, item in enumerate(line_items, start=r + 1):
        ws.cell(i, 1, item)
        for j, ton in enumerate(scales, start=2):
            v = result.opex.get(ton, {}).get(item, 0.0)
            ws.cell(i, j, v); _money_fmt(ws.cell(i, j))
    total_row = r + 1 + len(line_items)
    c = ws.cell(total_row, 1, "TOTAL OPEX"); c.font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(total_row, j, result.opex_total[ton])
        c.font = _bold(); _money_fmt(c)
    ws.column_dimensions["A"].width = 50
    for j in range(2, len(scales) + 2):
        ws.column_dimensions[chr(64 + j)].width = 22


def _write_revenue(wb, result, inp, db) -> None:
    ws = wb.create_sheet("Revenue_breakdown")
    r = _write_note_block(ws, [
        "Annual revenue per output product = (kg-product/y) × ($/kg market price).",
        "Market prices come from the Component DB (see Components sheet). Prices can be swept via price levers (see Levers and Price_* sheets).",
        "VERIFY: product mass/y on Flows_annual × price on this sheet = revenue. Sum of revenue lines = Summary's Revenue total.",
    ])
    scales = list(inp.scales_ton)
    products = sorted({k for ton in scales
                       for k in result.revenue.get(ton, {}).keys()})
    ws.cell(r, 1, "Product"); ws.cell(r, 1).font = _bold()
    ws.cell(r, 2, "Price ($/kg)"); ws.cell(r, 2).font = _bold()
    for j, ton in enumerate(scales, start=3):
        c = ws.cell(r, j, f"{ton} t-feed/batch ($/y)")
        c.font = _bold(); c.fill = _header_fill()
    for i, prod in enumerate(products, start=r + 1):
        ws.cell(i, 1, prod)
        try:
            price = db.get(prod).price_low
            ws.cell(i, 2, price); _money_fmt2(ws.cell(i, 2))
        except Exception:
            ws.cell(i, 2, "—")
        for j, ton in enumerate(scales, start=3):
            v = result.revenue.get(ton, {}).get(prod, 0.0)
            ws.cell(i, j, v); _money_fmt(ws.cell(i, j))
    total_row = r + 1 + len(products)
    c = ws.cell(total_row, 1, "TOTAL Revenue"); c.font = _bold()
    for j, ton in enumerate(scales, start=3):
        c = ws.cell(total_row, j, result.revenue_total[ton])
        c.font = _bold(); _money_fmt(c)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    for j in range(3, len(scales) + 3):
        ws.column_dimensions[chr(64 + j)].width = 24


def _write_flows(wb, result, inp) -> None:
    ws = wb.create_sheet("Flows_annual")
    r = _write_note_block(ws, [
        "Annual mass flow of every component (kg/y) at each scale.",
        "For inputs with recovery>0, this is the MAKEUP (= mass × (1-recovery) × batches/y), not the total reactor inventory.",
        "VERIFY: feedstock makeup × $/kg = OPEX 'Feedstock - X' line. Product mass × price = Revenue line. If these don't match, recovery or flow_mode is wrong.",
    ])
    scales = list(inp.scales_ton)
    components = sorted({k for ton in scales
                         for k in result.flows_annual_kg.get(ton, {}).keys()})
    ws.cell(r, 1, "Component"); ws.cell(r, 1).font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(r, j, f"{ton} t-feed/batch (kg/y)")
        c.font = _bold(); c.fill = _header_fill()
    for i, comp in enumerate(components, start=r + 1):
        ws.cell(i, 1, comp)
        for j, ton in enumerate(scales, start=2):
            v = result.flows_annual_kg.get(ton, {}).get(comp, 0.0)
            c = ws.cell(i, j, v); c.number_format = '#,##0'
    ws.column_dimensions["A"].width = 28
    for j in range(2, len(scales) + 2):
        ws.column_dimensions[chr(64 + j)].width = 22


def _write_levers(wb, build_fn) -> None:
    ws = wb.create_sheet("Levers")
    r = _write_note_block(ws, [
        "Every parameter that can be passed to build(...) as a keyword argument.",
        "LAB_DEFAULTS = what the current build() uses by default. COMMERCIAL_TARGETS = order-of-magnitude target for scale-up (used only by Gap_report).",
        "VERIFY: all your design assumptions are here. If a number isn't a lever, you can't sweep it — promote it to LAB_DEFAULTS in the builder.",
    ])
    mod = importlib.import_module(build_fn.__module__)
    lab = getattr(mod, "LAB_DEFAULTS", {})
    comm = getattr(mod, "COMMERCIAL_TARGETS", {})
    headers = ["Lever name", "Lab value", "Commercial target", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    notes = {
        "li_content_in_feed":       "Li wt% in dried black mass — supply-side quality",
        "li_recovery":              "fraction of feed Li recovered as Li2CO3",
        "ball_mill_energy_kwh_per_t": "ball-mill electricity per ton feed",
        "reagent_stoich_factor":    "NH4Cl excess over stoichiometric",
        "na2co3_stoich_factor":     "Na2CO3 excess over stoichiometric",
        "water_recovery":           "fraction of process water recycled",
        "fe_waste_disposal_usd_per_kg": "landfill cost — varies by region",
        "li2co3_price":             "market $/kg battery-grade",
        "lfp_feed_price":           "purchase $/kg from battery collectors",
        "nh4cl_price":              "industrial bulk $/kg",
        "na2co3_price":             "industrial bulk $/kg",
        "vanillin_selectivity":     "FRP thermal selectivity to vanillin",
        "acetovanillone_selectivity": "FRP thermal selectivity to AV",
        "lignin_to_h2":             "kg-lignin per kg-H2 from electron balance",
        "j_mA_per_cm2":             "electrolyzer current density",
        "fe_h2":                    "Faradaic efficiency for H2",
        "cell_voltage":             "V — drives electricity OPEX",
        "pma_recovery":             "PMA in-loop recycle fraction",
        "solvent_recovery_2methf":  "2-MeTHF distillation recovery",
    }
    for i, lever in enumerate(sorted(lab), start=r + 1):
        ws.cell(i, 1, lever)
        ws.cell(i, 2, lab[lever])
        ws.cell(i, 3, comm.get(lever, "—"))
        ws.cell(i, 4, notes.get(lever, ""))
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 60


def _write_gap(wb, build_fn) -> None:
    ws = wb.create_sheet("Gap_report")
    r = _write_note_block(ws, [
        "For each lever: lab value, commercial target, the single-lever 'breakeven alone' value, and economic impact of moving lab→commercial.",
        "Levers are sorted by |Δ impact| — top of list = biggest economic lever, where R&D investment matters most.",
        "VERIFY: top lever should match your engineering intuition. If a lever you expect to matter is at the bottom (or 'unreachable alone'), the model may not be capturing its effect.",
    ])
    mod = importlib.import_module(build_fn.__module__)
    lab = getattr(mod, "LAB_DEFAULTS", {})
    comm = getattr(mod, "COMMERCIAL_TARGETS", {})

    p, db, inp = build_fn()
    res = run_tea(p, db, inp)
    scale = max(inp.scales_ton)
    baseline = res.net_per_kg_feedstock.get(scale, float("nan"))

    ws.cell(r, 1, "Baseline (lab defaults)"); ws.cell(r, 1).font = _bold()
    ws.cell(r, 2, baseline); _money_fmt2(ws.cell(r, 2))
    r += 1
    ws.cell(r, 1, "Scale"); ws.cell(r, 1).font = _bold()
    ws.cell(r, 2, f"{scale} t-feed/batch")
    r += 2

    headers = ["Lever", "Lab value", "Commercial target",
               "Breakeven (alone)", "Metric @ commercial",
               "Δ impact ($/kg-feed)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    rows = []
    for lever, lab_val in lab.items():
        if not isinstance(lab_val, (int, float)):
            continue
        hi = max((comm.get(lever) or lab_val) * 50, lab_val * 50)
        lo = max(lab_val * 0.01, 1e-6)
        if lab_val < 1.0 and (comm.get(lever) or 0) < 1.0:
            hi = min(hi, 0.99)
        b_star = breakeven(build_fn, parameter=lever, target=0.0,
                           metric="net_per_kg_feedstock",
                           bracket=(lo, hi))
        if lever in comm:
            p2, db2, inp2 = build_fn(**{lever: comm[lever]})
            r2 = run_tea(p2, db2, inp2)
            at_comm = r2.net_per_kg_feedstock.get(scale, float("nan"))
            impact = at_comm - baseline
        else:
            at_comm, impact = None, None
        rows.append((lever, lab_val, comm.get(lever),
                     b_star, at_comm, impact))
    rows.sort(key=lambda x: (x[5] is None, -abs(x[5] or 0.0)))
    for i, (lev, lab_val, c_val, b_star, at_c, impact) in enumerate(rows, start=r + 1):
        ws.cell(i, 1, lev)
        ws.cell(i, 2, lab_val); _num_fmt(ws.cell(i, 2))
        ws.cell(i, 3, c_val if c_val is not None else "—")
        if c_val is not None:
            _num_fmt(ws.cell(i, 3))
        ws.cell(i, 4, b_star if b_star is not None else "unreachable alone")
        if b_star is not None:
            _num_fmt(ws.cell(i, 4))
        ws.cell(i, 5, at_c if at_c is not None else "—")
        if at_c is not None:
            _money_fmt2(ws.cell(i, 5))
        ws.cell(i, 6, impact if impact is not None else "—")
        if impact is not None:
            _money_fmt2(ws.cell(i, 6))
    ws.column_dimensions["A"].width = 32
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22


def _write_sweep(wb, build_fn, lever: str, values: List[float], inp) -> None:
    ws = wb.create_sheet(_safe_sheet_name(f"Sweep_{lever}"))
    r = _write_note_block(ws, [
        f"Sensitivity sweep on lever '{lever}'. Each row = one lever value; columns show the resulting $/kg-feed at each scale.",
        f"Last three columns (Revenue/OPEX/Net) at the LARGEST scale show what's actually changing.",
        "VERIFY: trend should make physical sense (e.g. higher recovery → better $/kg-feed). If a lever's curve is flat, the model isn't picking up its effect.",
    ])
    scales = list(inp.scales_ton)
    metric = "net_per_kg_feedstock"
    ws.cell(r, 1, "Sweep lever:"); ws.cell(r, 1).font = _bold()
    ws.cell(r, 2, lever)
    r += 1
    ws.cell(r, 1, "Metric:"); ws.cell(r, 1).font = _bold()
    ws.cell(r, 2, metric)
    r += 2

    ws.cell(r, 1, lever); ws.cell(r, 1).font = _bold()
    for j, ton in enumerate(scales, start=2):
        c = ws.cell(r, j, f"$/kg-feed @ {ton}t-feed/batch")
        c.font = _bold(); c.fill = _header_fill()
    biggest = max(scales)
    extra_headers = [f"Revenue ($/y) @ {biggest}t",
                     f"OPEX ($/y) @ {biggest}t",
                     f"Net profit ($/y) @ {biggest}t"]
    for k, eh in enumerate(extra_headers):
        c = ws.cell(r, len(scales) + 2 + k, eh)
        c.font = _bold(); c.fill = _header_fill()
    for i, v in enumerate(values, start=r + 1):
        ws.cell(i, 1, v); _num_fmt(ws.cell(i, 1))
        p, db, inp2 = build_fn(**{lever: v})
        res = run_tea(p, db, inp2)
        for j, ton in enumerate(scales, start=2):
            ws.cell(i, j, res.net_per_kg_feedstock.get(ton, 0.0))
            _money_fmt2(ws.cell(i, j))
        ws.cell(i, len(scales) + 2, res.revenue_total[biggest])
        _money_fmt(ws.cell(i, len(scales) + 2))
        ws.cell(i, len(scales) + 3, res.opex_total[biggest])
        _money_fmt(ws.cell(i, len(scales) + 3))
        ws.cell(i, len(scales) + 4, res.net_profit[biggest])
        _money_fmt(ws.cell(i, len(scales) + 4))
    ws.column_dimensions["A"].width = 18
    for col in range(2, len(scales) + 5):
        ws.column_dimensions[chr(64 + col)].width = 22


def _write_price_grid(wb, build_fn,
                      lever_x: str, values_x: List[float],
                      lever_y: str, values_y: List[float],
                      inp) -> None:
    ws = wb.create_sheet(_safe_sheet_name(f"Price_{lever_x}_{lever_y}"))
    r = _write_note_block(ws, [
        f"2-D price grid: rows = {lever_y} ($/kg), columns = {lever_x} ($/kg). Each cell = $/kg-feed at largest scale.",
        "Positive = profit per kg feedstock; negative = loss. Look for the breakeven contour (sign change) to see what price combinations make the process viable.",
        "VERIFY: profit should increase as product price rises and feed price falls. If it doesn't, a price lever isn't wired correctly in build().",
    ])
    scale = max(inp.scales_ton)
    ws.cell(r, 1, f"{lever_y} ↓ / {lever_x} →  ($/kg-feed @ {scale}t)")
    ws.cell(r, 1).font = _bold()
    for j, x in enumerate(values_x, start=2):
        c = ws.cell(r, j, x); c.font = _bold(); c.fill = _header_fill()
    for i, y in enumerate(values_y, start=r + 1):
        c = ws.cell(i, 1, y); c.font = _bold(); c.fill = _header_fill()
        for j, x in enumerate(values_x, start=2):
            p, db, inp2 = build_fn(**{lever_x: x, lever_y: y})
            res = run_tea(p, db, inp2)
            ws.cell(i, j, res.net_per_kg_feedstock.get(scale, 0.0))
            _money_fmt2(ws.cell(i, j))
    ws.column_dimensions["A"].width = 30
    for col in range(2, len(values_x) + 2):
        ws.column_dimensions[chr(64 + col)].width = 14


def _write_components(wb, db) -> None:
    ws = wb.create_sheet("Components")
    r = _write_note_block(ws, [
        "Every component referenced by the process: molecular weight, role (input/output/catalyst/etc.), market price, and source.",
        "These are the prices the engine actually uses for revenue and feedstock OPEX. Price levers in build() override these at run time.",
        "VERIFY: every input AND output stream should appear here. If you see a stream missing, the engine's revenue/OPEX for it will be $0.",
    ])
    headers = ["Name", "MW (g/mol)", "Role", "Price ($/kg)", "Source"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.font = _bold(); c.fill = _header_fill()
    components = []
    try:
        for name in sorted(db._components.keys()):
            components.append(db.get(name))
    except AttributeError:
        try:
            for c in db:
                components.append(c)
        except Exception:
            pass
    for i, c in enumerate(components, start=r + 1):
        ws.cell(i, 1, c.name)
        ws.cell(i, 2, c.mw)
        ws.cell(i, 3, c.role)
        ws.cell(i, 4, c.price_low or 0.0); _money_fmt2(ws.cell(i, 4))
        ws.cell(i, 5, getattr(c, "price_ref", "") or "")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 40


# ----------------------------------------------------------------------------
# Public API
# ----------------------------------------------------------------------------

def full_report(
    build_fn: Callable,
    output_path: str,
    sweep_specs: Optional[Sequence[Tuple[str, List[float]]]] = None,
    price_grid: Optional[Tuple[str, List[float], str, List[float]]] = None,
    pfd_png: Optional[str] = None,
) -> str:
    """Generate one comprehensive Excel workbook with PFD + tables + sweeps.

    Args:
        build_fn: parametric process builder
        output_path: where to write the .xlsx
        sweep_specs: list of (lever_name, [values]). None = auto-sweep every
                     numeric lever in LAB_DEFAULTS with default ±50% range.
        price_grid: optional (lever_x, values_x, lever_y, values_y) for a
                    2-D price sensitivity sheet.
        pfd_png: path to pre-rendered PFD PNG. If None, will auto-render.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl required. pip install openpyxl")

    _ensure_dir(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    p, db, inp = build_fn()
    res = run_tea(p, db, inp)

    # Render all 3 diagrams (auto if not provided)
    bfd_png = pid_png = None
    if pfd_png is None:
        try:
            from .pfd_renderer import render_bfd, render_pfd, render_pid
            tmpdir = tempfile.gettempdir()
            bfd_png = os.path.join(tmpdir, "_tea_bfd.png")
            pfd_png = os.path.join(tmpdir, "_tea_pfd.png")
            pid_png = os.path.join(tmpdir, "_tea_pid.png")
            render_bfd(p, bfd_png)
            render_pfd(p, pfd_png)
            render_pid(p, pid_png)
        except Exception as e:
            print(f"Diagram render skipped: {e}")
            bfd_png = pfd_png = pid_png = None

    # ---- Sheet order ----
    _write_readme(wb, build_fn, inp)
    if bfd_png and pfd_png and pid_png and all(
        os.path.exists(p) for p in [bfd_png, pfd_png, pid_png]):
        _write_diagrams(wb, p, bfd_png, pfd_png, pid_png)
    _write_process_structure(wb, p, db, inp)
    _write_summary(wb, build_fn, res, inp, db)
    _write_capex(wb, res, inp)
    _write_opex(wb, res, inp)
    _write_revenue(wb, res, inp, db)
    _write_flows(wb, res, inp)
    _write_levers(wb, build_fn)
    _write_gap(wb, build_fn)

    if sweep_specs is None:
        mod = importlib.import_module(build_fn.__module__)
        lab = getattr(mod, "LAB_DEFAULTS", {})
        sweep_specs = []
        for lever, lab_val in lab.items():
            if not isinstance(lab_val, (int, float)):
                continue
            if lab_val == 0:
                vals = [0.0, 0.01, 0.05, 0.1, 0.2, 0.5, 1.0]
            elif lab_val < 1.0:
                vals = [max(lab_val * 0.5, 1e-4),
                        lab_val * 0.75, lab_val * 0.9,
                        lab_val, min(lab_val * 1.1, 0.999),
                        min(lab_val * 1.25, 0.999),
                        min(lab_val * 1.5, 0.999)]
            else:
                vals = [lab_val * 0.5, lab_val * 0.75, lab_val * 0.9,
                        lab_val, lab_val * 1.25, lab_val * 1.5, lab_val * 2.0]
            sweep_specs.append((lever, sorted(set(vals))))
    for lever, values in sweep_specs:
        _write_sweep(wb, build_fn, lever, values, inp)

    if price_grid:
        lever_x, values_x, lever_y, values_y = price_grid
        _write_price_grid(wb, build_fn, lever_x, values_x,
                          lever_y, values_y, inp)

    _write_components(wb, db)

    wb.save(output_path)
    return output_path
