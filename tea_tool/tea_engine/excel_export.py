"""Write a TEA xlsx whose section layout mirrors the reference TEA Summary.

Keeps the same numbered headings (1-10) so the output is directly
diff-able with the paper's spreadsheet.
"""
from __future__ import annotations
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .components import ComponentDB
from .equipment import CEPCI
from .process import Process
from .tea import TEAInputs, TEAResult, sensitivity_one_param


_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_SECTION_FONT = Font(bold=True, size=12, color="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_BG = PatternFill("solid", fgColor="305496")
_BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                 right=Side(style="thin", color="BFBFBF"),
                 top=Side(style="thin", color="BFBFBF"),
                 bottom=Side(style="thin", color="BFBFBF"))


def _w(ws, row: int, col: int, val, *, bold=False, fill=None, font=None,
       fmt=None, border=False, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    if bold:
        c.font = Font(bold=True)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _BORDER
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c


def _section_header(ws, row: int, title: str) -> int:
    _w(ws, row, 2, title, font=_SECTION_FONT)
    return row + 2


def _table_header(ws, row: int, headers: List[str], col0: int = 2) -> None:
    for i, h in enumerate(headers):
        _w(ws, row, col0 + i, h, font=_HEADER_FONT, fill=_HEADER_BG, border=True, align="center")


def _price_lookup_urls(db: ComponentDB, name: str) -> List[str]:
    """First lookup URL(s) attached to a component via the YAML PriceDB,
    or [] when no PriceDB is wired into the ComponentDB."""
    pdb = getattr(db, "_pricedb", None)
    if pdb is None:
        return []
    try:
        return pdb.lookup_urls(name) or []
    except Exception:
        return []


def export_tea_xlsx(path: str, process: Process, db: ComponentDB,
                   inputs: TEAInputs, result: TEAResult,
                   sensitivity_specs: Dict[str, Dict] = None) -> None:
    wb = Workbook()

    # ===================== SHEET 1: TEA Summary =====================
    ws = wb.active
    ws.title = "TEA Summary"
    ws.column_dimensions["B"].width = 42
    for col in "CDEFG":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["H"].width = 50

    scales = list(inputs.scales_ton)
    scale_cols = [f"{int(s)} ton feed ($/y)" for s in scales]

    r = 2
    _w(ws, r, 2, f"{process.name} Techno-Economic Analysis (TEA) Summary", font=_TITLE_FONT)
    r += 1
    _w(ws, r, 2, process.description, font=Font(italic=True, color="595959"))
    r += 2

    # 1. Experimental Conditions & Scale-up
    r = _section_header(ws, r, "1. Experimental Conditions & Scale-up")
    _table_header(ws, r, ["Parameter", "Experimental (g)", "Scale-up (kg/batch)", "Annual (kg/y)"])
    r += 1
    largest = max(scales)
    sf_largest = largest * 1e6 / process.streams.inputs[0].mass_per_batch_g
    for s in process.streams.inputs:
        _w(ws, r, 2, f"{s.component} (Input)", border=True)
        _w(ws, r, 3, s.mass_per_batch_g, fmt="0.0000", border=True, align="right")
        _w(ws, r, 4, s.mass_per_batch_g * sf_largest / 1000, fmt="#,##0.0", border=True, align="right")
        _w(ws, r, 5, result.flows_annual_kg[largest][s.component], fmt="#,##0.0", border=True, align="right")
        r += 1
    for s in process.streams.outputs:
        _w(ws, r, 2, f"{s.component} (Output)", border=True)
        _w(ws, r, 3, s.mass_per_batch_g, fmt="0.0000", border=True, align="right")
        _w(ws, r, 4, s.mass_per_batch_g * sf_largest / 1000, fmt="#,##0.0", border=True, align="right")
        _w(ws, r, 5, result.flows_annual_kg[largest][s.component], fmt="#,##0.0", border=True, align="right")
        r += 1
    r += 1

    # 2. Key Assumptions
    r = _section_header(ws, r, "2. Key Assumptions")
    _table_header(ws, r, ["Parameter", "Value", "Reference"])
    r += 1
    rows = [
        ("Reaction time (1 Batch, h)", inputs.batch_hours,
         "User-supplied lab procedure"),
        ("Discount rate", inputs.discount_rate,
         "Joule 5, 2479-2503 (2021)"),
        ("Lifetime (years)", inputs.lifetime_years,
         "Joule 5, 2479-2503 (2021)"),
        ("Capital recovery factor", inputs.crf,
         "i(1+i)^n / ((1+i)^n − 1)"),
        ("Capacity factor", inputs.capacity_factor,
         "Nat Commun 12, 4679 (2021)"),
        ("Scaling factor (CAPEX)", 0.6,
         "Six-tenths rule (Peters & Timmerhaus 5e)"),
        ("CEPCI target", f"{inputs.cepci_target_year} ({CEPCI[inputs.cepci_target_year]})",
         "Chemical Engineering Plant Cost Index"),
        ("OSBL fraction of ISBL", inputs.osbl_fraction,
         "Towler & Sinnott, Chemical Engineering Design 2e"),
        ("Maintenance (% of CAPEX)", inputs.maintenance_fraction,
         "Towler & Sinnott (typical 5–10 %)"),
        ("Operation (% of CAPEX)", inputs.operation_fraction,
         "Towler & Sinnott (typical 5–10 %)"),
        ("Batches / year", inputs.batches_per_year,
         "365×24×CF / batch_h"),
    ]
    for s in process.streams.inputs:
        if s.recovery > 0:
            comp = db.get(s.component) if s.component in db else None
            ref = (comp.price_ref if comp and comp.price_ref else "lab procedure / engineering assumption")
            rows.append((f"{s.component} recovery", s.recovery, ref))
    for k, v in process.meta.items():
        if k.startswith("__"):
            continue  # private provenance entries (e.g. __matlab_sizing)
        if k.endswith("_$_per_ton_per_y"):
            ref = "Utility coefficient (annual $/y at 1-ton/batch baseline)"
        else:
            ref = "Process-specific parameter (lab data / engineering)"
        rows.append((f"meta: {k}", v, ref))
    for name, value, ref in rows:
        _w(ws, r, 2, name, border=True)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            fmt = "0.0000" if abs(value) < 1 else "#,##0.0000"
        else:
            fmt = None
        _w(ws, r, 3, value, fmt=fmt, border=True, align="right")
        if ref:
            _w(ws, r, 4, ref, border=True)
        r += 1
    r += 1

    # 3. CAPEX Summary
    r = _section_header(ws, r, "3. CAPEX Summary (Equipment + Installation)")
    _table_header(ws, r, ["Equipment Category"] + scale_cols)
    r += 1
    section_labels = list(result.capex_section[scales[0]].keys())
    for sec_label in section_labels:
        _w(ws, r, 2, sec_label, border=True)
        for i, sc in enumerate(scales):
            _w(ws, r, 3 + i, result.capex_section[sc][sec_label], fmt="#,##0", border=True, align="right")
        r += 1
    _w(ws, r, 2, "Total Equipment CAPEX", bold=True, border=True)
    for i, sc in enumerate(scales):
        _w(ws, r, 3 + i, result.capex_total[sc], fmt="#,##0", border=True, align="right", bold=True)
    r += 1
    _w(ws, r, 2, "Annualized Equipment CAPEX (×CRF)", border=True)
    for i, sc in enumerate(scales):
        _w(ws, r, 3 + i, result.capex_extra_annualized[sc]["Annualized Equipment CAPEX"],
           fmt="#,##0", border=True, align="right")
    r += 1
    # extra annualized lines (initial feedstock, distillation column...)
    extra_keys = [k for k in result.capex_extra_annualized[scales[0]] if k != "Annualized Equipment CAPEX"]
    for k in extra_keys:
        _w(ws, r, 2, f"Annualized CAPEX - {k}", border=True)
        for i, sc in enumerate(scales):
            _w(ws, r, 3 + i, result.capex_extra_annualized[sc][k], fmt="#,##0", border=True, align="right")
        r += 1
    _w(ws, r, 2, "Total Annualized CAPEX", bold=True, border=True, fill=_HEADER_FILL)
    for i, sc in enumerate(scales):
        _w(ws, r, 3 + i, result.capex_annualized[sc], fmt="#,##0", bold=True,
           border=True, fill=_HEADER_FILL, align="right")
    r += 2

    # 4. OPEX
    r = _section_header(ws, r, "4. OPEX Summary (Annual Operating Cost)")
    _table_header(ws, r, ["OPEX Category"] + scale_cols + ["Reference / Source", "Lookup URL"])
    r += 1
    line_keys = [k for k in result.opex[scales[0]].keys() if not k.startswith("__")]
    feedstock_components = {s.component for s in process.streams.inputs}
    for k in line_keys:
        _w(ws, r, 2, k, border=True)
        for i, sc in enumerate(scales):
            _w(ws, r, 3 + i, result.opex[sc][k], fmt="#,##0", border=True, align="right")
        # Pull source & URL from the matching feedstock component
        ref_text, ref_url = "", ""
        kk = k.strip()
        if kk.startswith("Feedstock - "):
            comp_name = kk.replace("Feedstock - ", "").replace(" makeup", "").strip()
            if comp_name in db:
                comp = db.get(comp_name)
                ref_text = comp.price_ref or ""
                urls = _price_lookup_urls(db, comp_name)
                ref_url = urls[0] if urls else ""
        _w(ws, r, 3 + len(scales), ref_text, border=True)
        _w(ws, r, 4 + len(scales), ref_url, border=True)
        r += 1
    for total_key in ("__Feedstock Total", "__Utility Total", "__Operation Total"):
        label = total_key.replace("__", "").replace(" Total", "") + " Total"
        _w(ws, r, 2, label, bold=True, border=True)
        for i, sc in enumerate(scales):
            _w(ws, r, 3 + i, result.opex[sc][total_key], fmt="#,##0", bold=True, border=True, align="right")
        r += 1
    _w(ws, r, 2, "Total OPEX", bold=True, border=True, fill=_HEADER_FILL)
    for i, sc in enumerate(scales):
        _w(ws, r, 3 + i, result.opex_total[sc], fmt="#,##0", bold=True,
           border=True, fill=_HEADER_FILL, align="right")
    r += 2

    # 5. Revenue
    r = _section_header(ws, r, "5. Revenue (Annual)")
    _table_header(ws, r, ["Product"] + scale_cols
                        + ["Unit Price ($/kg)", "Reference / Source", "Lookup URL"])
    r += 1
    for s in process.streams.outputs:
        comp = db.get(s.component)
        _w(ws, r, 2, s.component, border=True)
        for i, sc in enumerate(scales):
            _w(ws, r, 3 + i, result.revenue[sc][s.component], fmt="#,##0", border=True, align="right")
        _w(ws, r, 3 + len(scales), comp.price_low or 0.0, fmt="#,##0.000", border=True, align="right")
        _w(ws, r, 4 + len(scales), comp.price_ref or "", border=True)
        urls = _price_lookup_urls(db, s.component)
        _w(ws, r, 5 + len(scales), urls[0] if urls else "", border=True)
        r += 1
    _w(ws, r, 2, "Total Revenue", bold=True, border=True, fill=_HEADER_FILL)
    for i, sc in enumerate(scales):
        _w(ws, r, 3 + i, result.revenue_total[sc], fmt="#,##0", bold=True,
           border=True, fill=_HEADER_FILL, align="right")
    r += 2

    # 6. Profitability
    r = _section_header(ws, r, "6. Profitability Analysis")
    _table_header(ws, r, [""] + scale_cols)
    r += 1
    rows = [
        ("Total CAPEX (Annualized)", [result.capex_annualized[sc] for sc in scales]),
        ("Total OPEX",               [result.opex_total[sc] for sc in scales]),
        ("Total Cost",               [result.capex_annualized[sc] + result.opex_total[sc] for sc in scales]),
        ("Total Revenue",            [result.revenue_total[sc] for sc in scales]),
        ("Net Profit",               [result.net_profit[sc] for sc in scales]),
    ]
    for label, vals in rows:
        _w(ws, r, 2, label, bold=label in ("Net Profit", "Total Cost"), border=True)
        for i, v in enumerate(vals):
            _w(ws, r, 3 + i, v, fmt="#,##0",
               bold=label in ("Net Profit", "Total Cost"), border=True, align="right")
        r += 1
    r += 1

    # 7. MSP
    r = _section_header(ws, r, f"7. MSP (Minimum Selling Price of {inputs.msp_product})")
    _table_header(ws, r, ["Scale", f"MSP ($/kg {inputs.msp_product})"])
    r += 1
    for sc in scales:
        _w(ws, r, 2, f"{int(sc)} ton feed / batch", border=True)
        _w(ws, r, 3, result.msp[sc], fmt="0.0000", border=True, align="right")
        r += 1
    r += 1

    # 8. Cost / Revenue Breakdown (largest scale)
    r = _section_header(ws, r, f"8. Cost Breakdown ({int(largest)} ton feed)")
    _table_header(ws, r, ["Category", "Value ($/y)", "Portion (%)"])
    r += 1
    total = result.cost_breakdown["Total"]
    for k, v in result.cost_breakdown.items():
        if k == "Total":
            continue
        _w(ws, r, 2, k, border=True)
        _w(ws, r, 3, v, fmt="#,##0", border=True, align="right")
        _w(ws, r, 4, 100 * v / total if total else 0, fmt="0.00", border=True, align="right")
        r += 1
    _w(ws, r, 2, "Total", bold=True, border=True, fill=_HEADER_FILL)
    _w(ws, r, 3, total, fmt="#,##0", bold=True, border=True, fill=_HEADER_FILL, align="right")
    _w(ws, r, 4, 100.0, fmt="0.00", bold=True, border=True, fill=_HEADER_FILL, align="right")
    r += 2

    _w(ws, r, 2, f"Revenue Breakdown ({int(largest)} ton feed)", font=_SECTION_FONT)
    r += 2
    _table_header(ws, r, ["Product", "Value ($/y)", "Portion (%)"])
    r += 1
    rev_total = result.revenue_breakdown.get("Total", 0)
    for k, v in result.revenue_breakdown.items():
        if k == "Total":
            continue
        _w(ws, r, 2, k, border=True)
        _w(ws, r, 3, v, fmt="#,##0", border=True, align="right")
        _w(ws, r, 4, 100 * v / rev_total if rev_total else 0, fmt="0.00", border=True, align="right")
        r += 1
    r += 2

    # 9. Sensitivity (if any)
    if sensitivity_specs:
        r = _section_header(ws, r, f"9. Sensitivity Analysis ({int(largest)} ton feed)")
        for label, spec in sensitivity_specs.items():
            _table_header(ws, r, [label, f"MSP ($/kg {inputs.msp_product})"])
            r += 1
            data = sensitivity_one_param(process, db, inputs, spec["param"], spec["values"])
            for v, msp in data:
                _w(ws, r, 2, v, border=True, fmt="0.000000")
                _w(ws, r, 3, msp, fmt="0.0000", border=True, align="right")
                r += 1
            r += 1

    # ===================== SHEET 2: Streams & Equipment detail =====================
    ws2 = wb.create_sheet("Process Detail")
    ws2.column_dimensions["B"].width = 32
    for col in "CDEFGH":
        ws2.column_dimensions[col].width = 18

    r = 2
    _w(ws2, r, 2, "Material Flows", font=_SECTION_FONT)
    r += 2
    _table_header(ws2, r, ["Component", "Role", "Lab (g/batch)",
                           f"{int(largest)}-ton (kg/batch)", f"{int(largest)}-ton (kg/y)",
                           "Recovery", "Flow mode", "Init. charge (kg/ton)",
                           "Replace (months)"])
    r += 1
    for s in process.streams.inputs + process.streams.outputs:
        _w(ws2, r, 2, s.component, border=True)
        _w(ws2, r, 3, s.role, border=True)
        _w(ws2, r, 4, s.mass_per_batch_g, fmt="0.0000", border=True, align="right")
        _w(ws2, r, 5, result.flows_per_batch_kg[largest][s.component], fmt="#,##0.0", border=True, align="right")
        _w(ws2, r, 6, result.flows_annual_kg[largest][s.component], fmt="#,##0.0", border=True, align="right")
        _w(ws2, r, 7, s.recovery, fmt="0.0%", border=True, align="right")
        _w(ws2, r, 8, getattr(s, "flow_mode", "continuous"), border=True)
        _w(ws2, r, 9, getattr(s, "initial_charge_kg_per_ton", 0.0),
           fmt="#,##0.00", border=True, align="right")
        _w(ws2, r, 10, getattr(s, "replacement_interval_months", 0.0),
           fmt="#,##0", border=True, align="right")
        r += 1
    r += 2

    _w(ws2, r, 2, "Equipment List (Installed cost at largest scale)", font=_SECTION_FONT)
    r += 2
    _table_header(ws2, r, ["Section", "Equipment", "Base cost ($)",
                           "Install. factor", "CEPCI ref", "Scaling exp.",
                           f"Cost @ {int(largest)} ton ($)", "Lifetime (y)"])
    r += 1
    cepci_t = CEPCI[inputs.cepci_target_year]
    for eq in process.equipment.items:
        cost = eq.installed_cost(cepci_t, largest, process.meta)
        life = eq.lifetime_years
        life_disp = life if life is not None else "(plant)"
        for i, v in enumerate([eq.section, eq.name, eq.base_cost, eq.installation_factor,
                               eq.cepci_ref, eq.scaling_factor, cost, life_disp]):
            fmt = "#,##0" if i in (2, 6) else ("0.00" if i in (3, 5) else None)
            _w(ws2, r, 2 + i, v, fmt=fmt, border=True,
               align="right" if i >= 2 else "left")
        r += 1

    # ===================== SHEET 3: Time Profile =====================
    try:
        from .timeline import (material_timeline, cashflow_timeline,
                                stream_events, equipment_events)
        ws3 = wb.create_sheet("Time Profile")
        ws3.column_dimensions["B"].width = 8
        for col in "CDEFGH":
            ws3.column_dimensions[col].width = 18

        rr = 2
        _w(ws3, rr, 2, "Stream events (one_time / periodic)", font=_SECTION_FONT)
        rr += 2
        _table_header(ws3, rr, ["Month", "Year", "Kind", "Component", "$"])
        rr += 1
        for ev in sorted(stream_events(process, db, largest, inputs),
                         key=lambda e: e.month):
            _w(ws3, rr, 2, ev.month, border=True, align="right")
            _w(ws3, rr, 3, round(ev.month / 12, 2), fmt="0.00",
               border=True, align="right")
            _w(ws3, rr, 4, ev.kind.replace("_", " "), border=True)
            _w(ws3, rr, 5, ev.component, border=True)
            _w(ws3, rr, 6, ev.amount_usd, fmt="$#,##0", border=True, align="right")
            rr += 1
        rr += 2

        _w(ws3, rr, 2, "Equipment replacement events", font=_SECTION_FONT)
        rr += 2
        _table_header(ws3, rr, ["Month", "Year", "Item", "Section", "$"])
        rr += 1
        for ev in sorted(equipment_events(process, largest, inputs),
                         key=lambda e: e.month):
            _w(ws3, rr, 2, ev.month, border=True, align="right")
            _w(ws3, rr, 3, round(ev.month / 12, 2), fmt="0.00",
               border=True, align="right")
            _w(ws3, rr, 4, ev.label.replace("Replace — ", ""), border=True)
            _w(ws3, rr, 5, ev.section, border=True)
            _w(ws3, rr, 6, ev.amount_usd, fmt="$#,##0", border=True, align="right")
            rr += 1
        rr += 2

        _w(ws3, rr, 2, "Yearly cash flow (largest scale)", font=_SECTION_FONT)
        rr += 2
        _table_header(ws3, rr, ["Year", "CAPEX", "OPEX", "Revenue",
                                 "Stream events", "Equipment events", "Net", "Cumulative"])
        rr += 1
        cf = cashflow_timeline(process, db, largest, inputs,
                               result.opex[largest], result.revenue[largest],
                               result.capex_total[largest])
        cf["year_int"] = (cf["month"] // 12).astype(int)
        yearly = (cf.groupby("year_int", as_index=False)
                    .agg(capex=("capex", "sum"), opex=("opex", "sum"),
                         revenue=("revenue", "sum"),
                         stream_event=("stream_event", "sum"),
                         equipment_event=("equipment_event", "sum"),
                         net=("net", "sum")))
        yearly["cumulative"] = yearly["net"].cumsum()
        for _, row in yearly.iterrows():
            _w(ws3, rr, 2, int(row["year_int"]), border=True, align="right")
            for i, col in enumerate(["capex", "opex", "revenue", "stream_event",
                                      "equipment_event", "net", "cumulative"]):
                _w(ws3, rr, 3 + i, float(row[col]), fmt="$#,##0",
                   border=True, align="right")
            rr += 1
    except Exception:
        # Time-profile sheet is best-effort; never fail the whole export over it.
        pass

    # ===================== SHEET 4: References =====================
    ws4 = wb.create_sheet("References")
    ws4.column_dimensions["B"].width = 28
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 14
    ws4.column_dimensions["E"].width = 50
    ws4.column_dimensions["F"].width = 70

    rr = 2
    _w(ws4, rr, 2,
       f"{process.name} — References & Sources",
       font=_TITLE_FONT)
    rr += 1
    _w(ws4, rr, 2,
       "Note: Echemi / Alibaba / Made-in-China URLs return 403 to automated tools "
       "(Cloudflare bot protection) but open normally in a regular browser.",
       font=Font(italic=True, color="595959"))
    rr += 2

    # 1. Material / price references
    _w(ws4, rr, 2, "1. Material price references", font=_SECTION_FONT)
    rr += 2
    _table_header(ws4, rr, ["Component", "Role", "Price ($/kg)",
                             "Source / Reference", "Lookup URL(s)"])
    rr += 1
    seen_components = set()
    for s in process.streams.inputs + process.streams.outputs:
        if s.component in seen_components:
            continue
        seen_components.add(s.component)
        if s.component not in db:
            continue
        comp = db.get(s.component)
        urls = _price_lookup_urls(db, s.component)
        url_text = "\n".join(urls) if urls else ""
        _w(ws4, rr, 2, s.component, border=True)
        _w(ws4, rr, 3, s.role, border=True)
        _w(ws4, rr, 4, comp.price_low or 0.0, fmt="#,##0.000", border=True, align="right")
        _w(ws4, rr, 5, comp.price_ref or "", border=True)
        _w(ws4, rr, 6, url_text, border=True)
        if urls:
            ws4.row_dimensions[rr].height = max(15 * len(urls), 30)
        rr += 1
    rr += 2

    # 2. Process-knob references (TEAInputs + paper-derived constants)
    _w(ws4, rr, 2, "2. TEA framework references", font=_SECTION_FONT)
    rr += 2
    _table_header(ws4, rr, ["Parameter", "Value", "", "Source"])
    rr += 1
    framework_refs = [
        ("Discount rate",          inputs.discount_rate,
         "Singh et al., Joule 5, 2479-2503 (2021) — Techno-Economic, Life-Cycle "
         "Analysis of Enzymatic PET Recycling",
         "https://www.sciencedirect.com/science/article/pii/S2542435121003032"),
        ("Plant lifetime (years)", inputs.lifetime_years,
         "Singh et al., Joule 5, 2479-2503 (2021)",
         "https://www.sciencedirect.com/science/article/pii/S2542435121003032"),
        ("Capacity factor",        inputs.capacity_factor,
         "Liu et al., Nat. Commun. 12, 4679 (2021) — Electrocatalytic Upcycling "
         "of PET to Commodity Chemicals and H2 Fuel",
         "https://www.nature.com/articles/s41467-021-25048-x"),
        ("Scaling factor (CAPEX)", 0.6,
         "Six-tenths rule (Peters, Timmerhaus, West — Plant Design and Economics, 5e)",
         ""),
        ("CEPCI base year",        inputs.cepci_target_year,
         "Chemical Engineering Plant Cost Index (paywalled subscription)",
         "https://www.chemengonline.com/pci-home"),
        ("OSBL fraction of ISBL",  inputs.osbl_fraction,
         "Towler & Sinnott, Chemical Engineering Design 2e",
         ""),
        ("Maintenance (% of CAPEX)", inputs.maintenance_fraction,
         "Towler & Sinnott (typical 5–10 %)", ""),
        ("Operation (% of CAPEX)",   inputs.operation_fraction,
         "Towler & Sinnott (typical 5–10 %)", ""),
        ("Reference template",     "PET Depolymerization (PMA + Electrolysis)",
         "In-house reference: 260402 TEA summary.xlsx", ""),
    ]
    for name, val, src, url in framework_refs:
        _w(ws4, rr, 2, name, border=True)
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            fmt = "0.0000" if abs(val) < 1 else "#,##0.0000"
        else:
            fmt = None
        _w(ws4, rr, 3, val, fmt=fmt, border=True, align="right")
        _w(ws4, rr, 5, src, border=True)
        _w(ws4, rr, 6, url, border=True)
        rr += 1
    rr += 2

    # 3. Process meta references (whatever the builder put in process.meta)
    _w(ws4, rr, 2, "3. Process-specific meta parameters", font=_SECTION_FONT)
    rr += 2
    _table_header(ws4, rr, ["Parameter", "Value", "", "Note"])
    rr += 1
    for k, v in process.meta.items():
        if k.startswith("__"):
            continue  # private provenance entries (e.g. __matlab_sizing)
        _w(ws4, rr, 2, k, border=True)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            fmt = "0.0000" if abs(v) < 1 else "#,##0.0000"
        else:
            fmt = None
        _w(ws4, rr, 3, v, fmt=fmt, border=True, align="right")
        note = ""
        if k.endswith("_$_per_ton_per_y"):
            note = "Annual utility cost at 1-ton/batch baseline (linear in ton)"
        _w(ws4, rr, 5, note, border=True)
        rr += 1
    rr += 2

    # 4. Equipment costing convention
    _w(ws4, rr, 2, "4. Equipment costing convention", font=_SECTION_FONT)
    rr += 2
    _table_header(ws4, rr, ["Aspect", "", "", "Reference"])
    rr += 1
    eq_refs = [
        ("Installed-cost scaling",
         "Cost(cap) = base × (cap/cap_ref)^0.6 × (CEPCI_target/CEPCI_ref)",
         "", "Six-tenths rule"),
        ("CEPCI table",
         "2016: 541.7  2018: 603.1  2020: 596.2  2021: 708.0  2022: 816.0  "
         "2023: 800.8  2024: 800.0",
         "", "Chemical Engineering CEPCI publication"),
        ("Lifetime overrides",
         "Items with lifetime_years < plant life use their own CRF",
         "", "Towler & Sinnott; standard short-life equipment treatment"),
    ]
    for label, desc, _spacer, src in eq_refs:
        _w(ws4, rr, 2, label, border=True)
        _w(ws4, rr, 3, desc, border=True)
        _w(ws4, rr, 5, src, border=True)
        ws4.row_dimensions[rr].height = 30
        rr += 1

    wb.save(path)
