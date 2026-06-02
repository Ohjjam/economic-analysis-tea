"""Lab-to-commercial gap report.

Given a parametric build_fn whose module declares LAB_DEFAULTS (and ideally
COMMERCIAL_TARGETS), this module:
  1. Runs a sensitivity sweep on every lever
  2. Solves the single-lever breakeven for each (others held at lab default)
  3. Ranks levers by economic impact (how much $/kg-feed they buy when moved
     from lab to commercial)
  4. Writes a tidy Excel report + a markdown summary string

Typical use:
    from processes.yim_2022_modular_flow_lignin import build
    from tea_engine.gap_report import gap_report

    md = gap_report(build, excel_path="out/lignin_gap.xlsx")
    print(md)
"""
from __future__ import annotations
import importlib
import os
from typing import Callable, Dict, List, Optional, Tuple

from .tea import run_tea
from .solver import breakeven


def _get_module_constants(build_fn: Callable) -> Tuple[Dict, Dict]:
    """Pull LAB_DEFAULTS and COMMERCIAL_TARGETS from build_fn's module."""
    mod = importlib.import_module(build_fn.__module__)
    lab = getattr(mod, "LAB_DEFAULTS", {})
    comm = getattr(mod, "COMMERCIAL_TARGETS", {})
    return lab, comm


def _default_bracket(lever: str, lab_value: float,
                     comm_value: Optional[float]) -> Tuple[float, float]:
    """Heuristic bracket for breakeven search around lab/commercial range."""
    if comm_value is not None:
        hi = max(comm_value * 50, lab_value * 50)
    else:
        hi = lab_value * 500
    lo = max(lab_value * 0.01, 1e-6)
    # For levers like selectivity / recovery that live in [0, 1]:
    if lab_value < 1.0 and (comm_value is None or comm_value < 1.0):
        hi = min(hi, 0.99)
        lo = max(lo, 1e-5)
    return lo, hi


def gap_report(
    build_fn: Callable,
    metric: str = "net_per_kg_feedstock",
    target: float = 0.0,
    scale_ton: Optional[float] = None,
    excel_path: Optional[str] = None,
) -> str:
    """Run gap analysis on every lever in LAB_DEFAULTS, return a markdown report.

    Optionally writes a tidy Excel workbook with the full table.
    """
    lab, comm = _get_module_constants(build_fn)
    if not lab:
        return ("**No LAB_DEFAULTS found in builder module.**\n"
                "Add `LAB_DEFAULTS = {...}` at module top to enable gap report.")

    # Baseline at lab defaults
    p, db, inp = build_fn()
    r = run_tea(p, db, inp)
    scale = scale_ton if scale_ton is not None else max(inp.scales_ton)
    baseline = getattr(r, metric).get(scale, float("nan"))

    rows = []
    for lever, lab_val in lab.items():
        bracket = _default_bracket(lever, lab_val, comm.get(lever))
        # 1. Solve single-lever breakeven
        b_star = breakeven(build_fn, parameter=lever, target=target,
                           metric=metric, bracket=bracket, scale_ton=scale_ton)
        # 2. Economic impact at commercial target (if defined)
        if lever in comm:
            p2, db2, inp2 = build_fn(**{lever: comm[lever]})
            r2 = run_tea(p2, db2, inp2)
            at_comm = getattr(r2, metric).get(scale, float("nan"))
            impact = at_comm - baseline
        else:
            at_comm = None
            impact = None

        rows.append({
            "lever": lever,
            "lab_value": lab_val,
            "commercial_target": comm.get(lever),
            "breakeven_value": b_star,
            "metric_at_commercial": at_comm,
            "impact_vs_lab": impact,
        })

    # Sort by absolute impact (None last)
    rows.sort(key=lambda r: (r["impact_vs_lab"] is None,
                             -abs(r["impact_vs_lab"] or 0.0)))

    # Excel out
    if excel_path:
        _write_excel(rows, baseline, metric, scale, excel_path)

    # Markdown summary
    md = [
        f"## Lab-to-commercial gap report",
        f"",
        f"- Builder: `{build_fn.__module__}`",
        f"- Metric: `{metric}` at {scale} t-feed/batch scale",
        f"- Baseline (all levers at lab default): **{baseline:+.4f}**",
        f"- Target: {target}",
        f"",
        f"| Lever | Lab | Commercial | Breakeven (alone) | At commercial | Impact (Δ vs lab) |",
        f"|---|---|---|---|---|---|",
    ]
    for row in rows:
        comm_str = (f"{row['commercial_target']}"
                    if row['commercial_target'] is not None else "—")
        bk = (f"{row['breakeven_value']:.4f}"
              if row['breakeven_value'] is not None else "unreachable alone")
        at_c = (f"{row['metric_at_commercial']:+.4f}"
                if row['metric_at_commercial'] is not None else "—")
        impact = (f"{row['impact_vs_lab']:+.4f}"
                  if row['impact_vs_lab'] is not None else "—")
        md.append(f"| `{row['lever']}` | {row['lab_value']} | {comm_str} | {bk} | {at_c} | {impact} |")
    md.append("")
    md.append("**Reading:** Levers near the top have the biggest economic impact.")
    md.append("'Breakeven alone' = unreachable means that lever can't reach target")
    md.append("while OTHER levers remain at lab values — you need to combine multiple.")
    return "\n".join(md)


def _write_excel(rows, baseline, metric, scale, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl required. pip install openpyxl")

    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "gap_report"

    # Header rows: baseline info
    ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=1, column=2, value=metric)
    ws.cell(row=2, column=1, value="Scale (t-feed/batch)").font = Font(bold=True)
    ws.cell(row=2, column=2, value=scale)
    ws.cell(row=3, column=1, value="Baseline @ lab defaults").font = Font(bold=True)
    ws.cell(row=3, column=2, value=baseline)

    # Table header
    header = ["Lever", "Lab value", "Commercial target",
              "Breakeven (alone)", "Metric @ commercial",
              "Δ vs lab (impact)"]
    for j, h in enumerate(header, start=1):
        c = ws.cell(row=5, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")

    # Body rows
    for i, r in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=r["lever"])
        ws.cell(row=i, column=2, value=r["lab_value"])
        ws.cell(row=i, column=3, value=r["commercial_target"])
        ws.cell(row=i, column=4, value=r["breakeven_value"])
        ws.cell(row=i, column=5, value=r["metric_at_commercial"])
        ws.cell(row=i, column=6, value=r["impact_vs_lab"])

    # Auto-width-ish (set fixed reasonable widths)
    widths = [28, 14, 18, 22, 22, 22]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    wb.save(path)
