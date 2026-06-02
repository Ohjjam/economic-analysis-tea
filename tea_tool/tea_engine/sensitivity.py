"""Sensitivity sweep API for forward-design mode.

Given a parametric `build_fn(**overrides) -> (process, db, inp)`, run TEA over
a range of values for one (or two) levers and return a tidy result.

Typical use:
    from processes.yim_2022_modular_flow_lignin import build
    from tea_engine.sensitivity import sweep

    s = sweep(build, "j_mA_per_cm2", [20, 50, 100, 200, 500, 1000],
              excel_path="out/lignin_j_sweep.xlsx",
              png_path="out/lignin_j_sweep.png")
    # s = {20: -4.78, 50: -2.96, 100: -1.51, 200: -0.91, 500: -0.79, 1000: -0.75}

Outputs:
    - Python dict (always)
    - Excel workbook with all scales × all values (if excel_path)
    - Matplotlib PNG plot (if png_path)
"""
from __future__ import annotations
import os
from typing import Callable, Dict, List, Optional

from .tea import run_tea


def sweep(
    build_fn: Callable,
    parameter: str,
    values: List[float],
    metric: str = "net_per_kg_feedstock",
    scale_ton: Optional[float] = None,
    excel_path: Optional[str] = None,
    png_path: Optional[str] = None,
    verbose: bool = False,
) -> Dict[float, float]:
    """Sweep one lever, single-scale headline.

    Args:
        build_fn: parametric process builder (e.g. lignin's build)
        parameter: lever name (must match build_fn's kwarg)
        values: list of lever values to evaluate
        metric: TEAResult field name to extract (e.g. "net_per_kg_feedstock",
                "msp", "net_profit", "capex_total", "opex_total")
        scale_ton: which scale point to report (default: largest in scales_ton)
        excel_path: write Excel workbook if set
        png_path: write matplotlib PNG if set

    Returns:
        {parameter_value: metric_at_that_value}
    """
    headline = {}
    all_scales: Dict[float, Dict[float, float]] = {}  # {value: {scale: metric}}

    for v in values:
        p, db, inp = build_fn(**{parameter: v})
        r = run_tea(p, db, inp)
        metric_dict = getattr(r, metric)
        all_scales[v] = dict(metric_dict)
        scale = scale_ton if scale_ton is not None else max(inp.scales_ton)
        headline[v] = metric_dict.get(scale, float("nan"))
        if verbose:
            print(f"  {parameter}={v}: {metric}@{scale}t = {headline[v]:+.4f}")

    if excel_path:
        _write_excel(all_scales, parameter, metric, excel_path)
    if png_path:
        _write_png(all_scales, parameter, metric, png_path)

    return headline


def sweep_2d(
    build_fn: Callable,
    parameter_x: str,
    values_x: List[float],
    parameter_y: str,
    values_y: List[float],
    metric: str = "net_per_kg_feedstock",
    scale_ton: Optional[float] = None,
    excel_path: Optional[str] = None,
    png_path: Optional[str] = None,
) -> Dict[tuple, float]:
    """Two-lever grid sweep. Returns {(x, y): metric}. Useful for
    "j vs selectivity" tradeoff plots.
    """
    grid: Dict[tuple, float] = {}
    for x in values_x:
        for y in values_y:
            p, db, inp = build_fn(**{parameter_x: x, parameter_y: y})
            r = run_tea(p, db, inp)
            scale = scale_ton if scale_ton is not None else max(inp.scales_ton)
            metric_dict = getattr(r, metric)
            grid[(x, y)] = metric_dict.get(scale, float("nan"))

    if excel_path:
        _write_excel_2d(grid, parameter_x, values_x,
                        parameter_y, values_y, metric, excel_path)
    if png_path:
        _write_png_2d(grid, parameter_x, values_x,
                      parameter_y, values_y, metric, png_path)

    return grid


# ----------------------------------------------------------------------------
# Excel + PNG writers (kept simple, single-purpose)
# ----------------------------------------------------------------------------

def _ensure_dir(path: str) -> None:
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def _write_excel(all_scales: Dict[float, Dict[float, float]],
                 parameter: str, metric: str, path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl required for Excel output. pip install openpyxl")

    _ensure_dir(path)
    wb = Workbook()
    ws = wb.active
    ws.title = f"sweep_{parameter}"[:31]

    # Header: parameter value | scale1 | scale2 | ... | scale_largest
    scales = sorted({s for sd in all_scales.values() for s in sd.keys()})
    ws.cell(row=1, column=1, value=parameter)
    for j, sc in enumerate(scales, start=2):
        ws.cell(row=1, column=j, value=f"{metric}@{sc}t-feed/batch")

    # Body
    for i, v in enumerate(sorted(all_scales), start=2):
        ws.cell(row=i, column=1, value=v)
        for j, sc in enumerate(scales, start=2):
            ws.cell(row=i, column=j, value=all_scales[v].get(sc))

    wb.save(path)


def _write_png(all_scales: Dict[float, Dict[float, float]],
               parameter: str, metric: str, path: str) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        raise RuntimeError("matplotlib required for PNG output. pip install matplotlib")

    _ensure_dir(path)
    fig, ax = plt.subplots(figsize=(8, 5))
    xs = sorted(all_scales.keys())
    scales = sorted({s for sd in all_scales.values() for s in sd.keys()})
    for sc in scales:
        ys = [all_scales[v].get(sc, float("nan")) for v in xs]
        ax.plot(xs, ys, marker="o", label=f"{sc}t-feed/batch")
    ax.axhline(0, color="gray", linewidth=0.5, linestyle="--")
    ax.set_xlabel(parameter)
    ax.set_ylabel(metric)
    ax.set_title(f"Sensitivity: {metric} vs {parameter}")
    ax.legend(title="scale", fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)


def _write_excel_2d(grid, px, vx, py, vy, metric, path):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl required for Excel output. pip install openpyxl")

    _ensure_dir(path)
    wb = Workbook()
    ws = wb.active
    ws.title = f"2d_{px[:10]}_{py[:10]}"[:31]
    ws.cell(row=1, column=1, value=f"{px} \\ {py}")
    for j, y in enumerate(vy, start=2):
        ws.cell(row=1, column=j, value=y)
    for i, x in enumerate(vx, start=2):
        ws.cell(row=i, column=1, value=x)
        for j, y in enumerate(vy, start=2):
            ws.cell(row=i, column=j, value=grid.get((x, y)))
    wb.save(path)


def _write_png_2d(grid, px, vx, py, vy, metric, path):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        raise RuntimeError("matplotlib+numpy required for PNG output.")

    _ensure_dir(path)
    Z = np.array([[grid.get((x, y), float("nan")) for y in vy] for x in vx])
    fig, ax = plt.subplots(figsize=(8, 6))
    im = ax.imshow(Z, aspect="auto", origin="lower",
                   extent=[min(vy), max(vy), min(vx), max(vx)],
                   cmap="RdYlGn")
    # Breakeven contour (metric=0)
    try:
        X, Y = np.meshgrid(vy, vx)
        cs = ax.contour(X, Y, Z, levels=[0.0], colors="black", linewidths=2)
        ax.clabel(cs, fmt={0.0: "breakeven"}, inline=True, fontsize=9)
    except Exception:
        pass
    fig.colorbar(im, ax=ax, label=metric)
    ax.set_xlabel(py)
    ax.set_ylabel(px)
    ax.set_title(f"2D sensitivity: {metric}")
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
